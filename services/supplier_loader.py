import os
import unicodedata

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_FILE = os.path.join(BASE_DIR, "Fournisseur Yoomi.xlsx")
IMAGE_FOLDER = os.path.join(BASE_DIR, "images")


def normalize_name(name):

    return (
        str(name)
        .lower()
        .replace(" ", "_")
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ç", "c")
        .replace("/", "")
    )


def normalize_header(value):

    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.lower().split())


def find_column(columns, *keywords):

    for col in columns:

        header = normalize_header(col)

        if all(keyword in header for keyword in keywords):
            return col

    return None


def find_product_column_index(ws):

    for cell in ws[1]:

        header = normalize_header(cell.value)

        if "nom" in header and "produit" in header:
            return cell.col_idx

    return 2


def extract_images_once(force=False):

    if not os.path.exists(EXCEL_FILE):
        print(f"Supplier file not found: {EXCEL_FILE}")
        return

    if not os.path.exists(IMAGE_FOLDER):
        os.makedirs(IMAGE_FOLDER)

    print("Checking product images...")

    wb = load_workbook(
        EXCEL_FILE,
        data_only=True
    )

    for ws in wb.worksheets:

        if not hasattr(ws, "_images") or not ws._images:
            continue

        product_column = find_product_column_index(ws)

        for image in ws._images:

            try:
                row = image.anchor._from.row + 1
                product_name = ws.cell(
                    row=row,
                    column=product_column
                ).value
            except Exception as exc:
                print("Image skip:", exc)
                continue

            if not product_name or not str(product_name).strip():
                continue

            filename = normalize_name(product_name)

            path = os.path.join(
                IMAGE_FOLDER,
                f"{filename}.png"
            )

            if os.path.exists(path) and not force:
                continue

            try:

                data = image._data()

                with open(path, "wb") as f:
                    f.write(data)

                print("Saved:", ascii(filename))

            except Exception as exc:

                print("Image skip:", exc)


def load_supplier_catalog():

    suppliers = {}

    print("Loading suppliers...")

    if not os.path.exists(EXCEL_FILE):
        print(f"Supplier file not found: {EXCEL_FILE}")
        return suppliers

    try:
        xls = pd.ExcelFile(EXCEL_FILE)
    except Exception as exc:
        print(f"Failed to load supplier Excel file: {exc}")
        return suppliers

    for sheet in xls.sheet_names:

        try:
            df = pd.read_excel(
                xls,
                sheet
            )
        except Exception as exc:
            print(f"Skipping sheet '{sheet}': {exc}")
            continue

        product_col = find_column(df.columns, "nom")

        if not product_col:
            continue

        unit_col = find_column(df.columns, "unite")
        items = []
        seen = set()

        for _, row in df.iterrows():

            raw_product = row.get(product_col)

            if pd.isna(raw_product) or not str(raw_product).strip():
                continue

            product = str(raw_product)

            if product in seen:
                continue

            seen.add(product)

            unit = ""
            if unit_col:
                raw_unit = row.get(unit_col)
                if pd.notna(raw_unit):
                    unit = str(raw_unit).strip()

            items.append(
                {
                    "name": product,
                    "unit": unit or "Pièce"
                }
            )

        if items:
            suppliers[sheet] = items

    return suppliers


def load_suppliers():

    return {
        supplier: [item["name"] for item in items]
        for supplier, items in load_supplier_catalog().items()
    }